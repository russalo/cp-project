"""
Ingest masters from the CP EWO Excel workbook.

Loads four sheets:
  - "Caltrans Rates 26-27" → CaltransSchedule + CaltransRateLine (factor-based)
  - "Union Rates"          → TradeClassification + LaborRate (effective today)
  - "Labor"                → Employee (named rows only; generic trades already
                             covered by Union Rates)
  - "Equipment"            → EquipmentType (own rates + CT match quality)

Idempotent by natural key. Re-running updates existing rows instead of
duplicating them. Does not touch EWO / WorkDay / line data.

Equipment rows: ``rate_reg`` is the value from the Equipment sheet's Rate
column (authoritative for CP). ``rate_ot`` and ``rate_standby`` are left
at 0 when there is no Caltrans link — populate them via admin, or a
follow-up pass that matches Equipment.CT Code against CaltransRateLine.

Usage:
    python manage.py ingest_ewo_workbook path/to/EWOexample.xlsx [--dry-run]
"""

from datetime import date
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from resources.models import (
    CaltransRateLine,
    CaltransSchedule,
    Employee,
    EquipmentType,
    LaborRate,
    TradeClassification,
)


# Best-effort mapping of trade names to union locals.
# Extend/correct in the admin after ingest.
TRADE_UNION_MAP = {
    'Operator': ('Operating Engineers Local 12', 'IUOE'),
    'Operator Foreman': ('Operating Engineers Local 12', 'IUOE'),
    'Laborer': ('Laborers Local 1184', 'LIUNA'),
    'Laborer Pipelayer': ('Laborers Local 1184', 'LIUNA'),
    'Laborer Foreman': ('Laborers Local 1184', 'LIUNA'),
    'Cement Mason': ('Cement Masons Local 600', 'OPCMIA'),
    'CM Foreman': ('Cement Masons Local 600', 'OPCMIA'),
    'Teamster': ('Teamsters Local 166', 'IBT'),
    'Mechanic': ('Operating Engineers Local 12', 'IUOE'),
    'Welder': ('N/A', 'NON'),
    'Superintendent': ('N/A', 'NON'),
}

# Equipment categories that typically do NOT burn fuel (override in admin if wrong).
# Everything else defaults to fuel_surcharge_eligible=True.
NON_FUEL_CATEGORIES = {
    'TRAILERS', 'SHORING', 'SAFETY', 'TRAFFIC CONTROL', 'HOSES', 'MISC',
}

CT_MATCH_NORMALIZE = {
    'exact': EquipmentType.CtMatchQuality.EXACT,
    'close': EquipmentType.CtMatchQuality.CLOSE,
    'none': EquipmentType.CtMatchQuality.NONE,
    '': EquipmentType.CtMatchQuality.NONE,
    None: EquipmentType.CtMatchQuality.NONE,
}


class Command(BaseCommand):
    help = 'Ingest EWO masters (Caltrans, Union Rates, Labor, Equipment) from the CP Excel workbook.'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str, help='Path to the .xlsx workbook')
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Parse and report counts without writing to the database.',
        )
        parser.add_argument(
            '--effective-date', type=str, default=None,
            help='Effective date for new LaborRate rows (ISO YYYY-MM-DD). Defaults to today.',
        )

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError as e:
            raise CommandError(
                'openpyxl is not installed. Add it to requirements or `pip install openpyxl`.'
            ) from e

        xlsx_path = Path(opts['xlsx_path'])
        if not xlsx_path.exists():
            raise CommandError(f'Workbook not found: {xlsx_path}')

        effective_date = (
            date.fromisoformat(opts['effective_date']) if opts['effective_date'] else date.today()
        )
        dry_run = opts['dry_run']

        self.stdout.write(self.style.NOTICE(f'Loading workbook: {xlsx_path}'))
        self.stdout.write(f'  effective_date: {effective_date}')
        self.stdout.write(f'  dry_run: {dry_run}')
        self.stdout.write('')

        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

        counts = {}

        with transaction.atomic():
            counts['caltrans'] = self._ingest_caltrans(wb, xlsx_path.name)
            counts['trades_and_rates'] = self._ingest_union_rates(wb, effective_date)
            counts['employees'] = self._ingest_labor_employees(wb)
            counts['equipment'] = self._ingest_equipment(wb)

            if dry_run:
                self.stdout.write(self.style.WARNING('\n--dry-run: rolling back all writes.'))
                transaction.set_rollback(True)

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('Summary:'))
        for key, stats in counts.items():
            self.stdout.write(f'  {key}: {stats}')

    # ── Caltrans ──────────────────────────────────────────────────────────────

    def _ingest_caltrans(self, wb, source_filename):
        """Caltrans Rates 26-27: Class Code | Make | Model/Description | Hourly Rate | RW Delay | OT Factor | Equipment Type."""
        ws = wb['Caltrans Rates 26-27']

        # Schedule window inferred from the sheet title; harcoded 2026-2027 for now.
        schedule, _ = CaltransSchedule.objects.update_or_create(
            schedule_year='2026-2027',
            defaults={
                'effective_date': date(2026, 4, 1),
                'expiry_date': date(2027, 3, 31),
                'source_file': source_filename,
            },
        )

        created, updated = 0, 0
        seen = set()
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i < 4:  # header rows are 1-3
                continue
            class_code, make, model_desc, hourly_rate, rw_delay, ot_factor, equip_type_desc = row[:7]
            if not class_code or hourly_rate is None:
                continue

            # make_code/make_desc: use the Make column for both (no separate code)
            make_code = _short_code(make, max_len=20)
            model_code = _short_code(model_desc, max_len=50)

            natural_key = (schedule.pk, class_code, make_code, model_code)
            if natural_key in seen:
                continue  # duplicate row in sheet — skip
            seen.add(natural_key)

            obj, was_created = CaltransRateLine.objects.update_or_create(
                schedule=schedule,
                class_code=class_code,
                make_code=make_code,
                model_code=model_code,
                defaults={
                    'class_desc': (equip_type_desc or '')[:200],
                    'make_desc': (str(make) or '')[:200],
                    'model_desc': (str(model_desc) or '')[:200],
                    'rental_rate': _as_decimal(hourly_rate, '0.00'),
                    'rw_delay_factor': _as_decimal(rw_delay, '0.0000'),
                    'ot_factor': _as_decimal(ot_factor, '0.0000'),
                    'unit': 'HR',
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(f'  Caltrans: schedule={schedule.schedule_year}, +{created} created, ~{updated} updated')
        return {'schedule': schedule.schedule_year, 'created': created, 'updated': updated}

    # ── Union Rates → TradeClassification + LaborRate ─────────────────────────

    def _ingest_union_rates(self, wb, effective_date):
        """Union Rates: Trade | REG Rate | OT Rate | DT Rate | Source."""
        ws = wb['Union Rates']

        created_trades, updated_trades = 0, 0
        created_rates, updated_rates = 0, 0
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 1:  # header
                continue
            trade_name, reg, ot, dt, source = row[:5]
            if not trade_name or reg is None:
                continue

            if trade_name not in TRADE_UNION_MAP:
                raise CommandError(
                    f'Unknown trade "{trade_name}" encountered in Union Rates sheet. '
                    f'Add it to TRADE_UNION_MAP in '
                    f'resources/management/commands/ingest_ewo_workbook.py '
                    f'(supply union_name and union_abbrev) before re-running.'
                )
            union_name, union_abbrev = TRADE_UNION_MAP[trade_name]
            trade, was_created = TradeClassification.objects.update_or_create(
                name=trade_name,
                defaults={'union_name': union_name, 'union_abbrev': union_abbrev},
            )
            if was_created:
                created_trades += 1
            else:
                updated_trades += 1

            # LaborRate: upsert by (trade, effective_date)
            _, was_created = LaborRate.objects.update_or_create(
                trade_classification=trade,
                effective_date=effective_date,
                defaults={
                    'rate_reg': _as_decimal(reg),
                    'rate_ot': _as_decimal(ot),
                    'rate_dt': _as_decimal(dt),
                    'notes': str(source or '')[:200],
                },
            )
            if was_created:
                created_rates += 1
            else:
                updated_rates += 1

        self.stdout.write(
            f'  Union Rates: trades +{created_trades} / ~{updated_trades}, '
            f'rates +{created_rates} / ~{updated_rates} (effective {effective_date})'
        )
        return {
            'trades_created': created_trades, 'trades_updated': updated_trades,
            'rates_created': created_rates, 'rates_updated': updated_rates,
        }

    # ── Labor sheet → Employee (named rows only) ──────────────────────────────

    def _ingest_labor_employees(self, wb):
        """Labor sheet: Code | Name | Trade | REG Rate | OT Rate | DT Rate | Notes.

        Rows 2–12 are generic trade codes (single letters / short codes like
        S, C, CF, L, LP, LF, O, OF, M, T, W) — skip those; trades were
        populated by Union Rates ingest. Rows 13+ are named employees whose
        Code follows ``[TRADE]-[LLFF]`` format.
        """
        ws = wb['Labor']
        # Pre-fetch trades into a dict — tiny table, avoids a query per employee.
        trade_by_name = {t.name: t for t in TradeClassification.objects.all()}
        created, updated, skipped = 0, 0, 0
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 1:
                continue
            code, name, trade_name, reg, ot, dt, notes = row[:7]
            if not code or not name:
                continue
            # Skip generic-trade rows (code has no hyphen, e.g. "L" or "OF")
            if '-' not in str(code):
                skipped += 1
                continue

            trade = trade_by_name.get(trade_name)
            if trade is None:
                skipped += 1
                self.stdout.write(self.style.WARNING(
                    f'    skip employee {code!r}: trade {trade_name!r} not found'
                ))
                continue

            _, was_created = Employee.objects.update_or_create(
                code=code,
                defaults={
                    'full_name': name,
                    'trade_classification': trade,
                    'active': True,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(
            f'  Employees: +{created} created, ~{updated} updated, {skipped} skipped (generic/unmapped)'
        )
        return {'created': created, 'updated': updated, 'skipped': skipped}

    # ── Equipment sheet → EquipmentType ───────────────────────────────────────

    def _ingest_equipment(self, wb):
        """Equipment: Code | Description | Category | Rate | Unit | CT Code | CT Match | Notes."""
        ws = wb['Equipment']
        created, updated = 0, 0
        match_counts = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 1:
                continue
            code, desc, category, rate, unit_, ct_code, ct_match, notes = row[:8]
            if not code or not desc:
                continue

            match_key = (ct_match or '').strip().lower()
            # Distinguish "CT code retired" entries (via notes) from plain None
            match_quality = CT_MATCH_NORMALIZE.get(match_key, EquipmentType.CtMatchQuality.NONE)
            if (notes or '').lower().startswith('ct code retired'):
                match_quality = EquipmentType.CtMatchQuality.RETIRED

            fuel_eligible = (category or '').strip().upper() not in NON_FUEL_CATEGORIES

            rate_reg = _as_decimal(rate, '0.00')

            # Use Equipment sheet "Code" as the unique name; description goes in notes if different.
            # This matches the CP code convention (e.g. "DT-4AX") and keeps the admin searchable.
            composed_notes = '\n'.join(
                filter(None, [
                    f'Description: {desc}' if desc else None,
                    f'Category: {category}' if category else None,
                    f'CT Code: {ct_code}' if ct_code else None,
                    f'Unit: {unit_}' if unit_ else None,
                    f'Note: {notes}' if notes else None,
                ])
            )

            # Sheet-sourced fields are authoritative on every run. Fields
            # populated later by admin or a follow-up linker (rate_ot,
            # rate_standby, caltrans_rate_line) are only set on CREATE so
            # re-running ingest never clobbers manual adjustments.
            sheet_defaults = {
                'rate_reg': rate_reg,
                'fuel_surcharge_eligible': fuel_eligible,
                'ct_match_quality': match_quality,
                'notes': composed_notes,
                'active': True,
            }
            et, was_created = EquipmentType.objects.get_or_create(
                name=code,  # natural key: the CP equipment code
                defaults={
                    **sheet_defaults,
                    'rate_ot': Decimal('0'),       # admin/follow-up fills; never touched on update
                    'rate_standby': Decimal('0'),  # admin/follow-up fills; never touched on update
                    'caltrans_rate_line': None,    # admin/follow-up links; never touched on update
                },
            )
            if was_created:
                created += 1
            else:
                # Update only the sheet-sourced fields.
                for k, v in sheet_defaults.items():
                    setattr(et, k, v)
                et.save(update_fields=list(sheet_defaults.keys()))
                updated += 1
            match_counts[match_quality] = match_counts.get(match_quality, 0) + 1

        match_summary = ', '.join(f'{k}={v}' for k, v in sorted(match_counts.items()))
        self.stdout.write(
            f'  Equipment: +{created} created, ~{updated} updated   ({match_summary})'
        )
        return {'created': created, 'updated': updated, 'by_match_quality': match_counts}


# ─── helpers ──────────────────────────────────────────────────────────────────


def _as_decimal(value, default='0'):
    """Coerce a spreadsheet cell value to a Decimal safely."""
    if value is None or value == '':
        return Decimal(default)
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _short_code(value, max_len=20):
    """Strip/normalize a string for use as a 'code'-style natural-key slug."""
    if value is None:
        return ''
    return str(value).strip()[:max_len]
